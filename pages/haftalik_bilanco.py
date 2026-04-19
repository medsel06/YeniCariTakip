"""Cari Takip - Haftalik Bilanco Sayfasi (Uretim Maliyet Takibi)"""
from datetime import date, datetime, timedelta
from nicegui import ui
from layout import create_layout, fmt_para, notify_ok, notify_err
from services.settings_service import get_company_settings
from services.stok_service import get_urun_list
from db import get_db

AY_ISIMLERI = {
    1: 'Ocak', 2: 'Şubat', 3: 'Mart', 4: 'Nisan', 5: 'Mayıs', 6: 'Haziran',
    7: 'Temmuz', 8: 'Ağustos', 9: 'Eylül', 10: 'Ekim', 11: 'Kasım', 12: 'Aralık',
}
AY_KISA = {
    1: 'Oca', 2: 'Şub', 3: 'Mar', 4: 'Nis', 5: 'May', 6: 'Haz',
    7: 'Tem', 8: 'Ağu', 9: 'Eyl', 10: 'Eki', 11: 'Kas', 12: 'Ara',
}


def _get_haftalar(yil, ay):
    """Secilen aydaki ISO haftalarini dondur. iso_yil ile birlikte."""
    from calendar import monthrange
    result = []
    seen = set()
    _, gun_sayisi = monthrange(yil, ay)
    for gun in range(1, gun_sayisi + 1):
        d = date(yil, ay, gun)
        iso_yil, iso_hafta, _ = d.isocalendar()
        key = (iso_yil, iso_hafta)
        if key in seen:
            continue
        seen.add(key)
        pazartesi = d - timedelta(days=d.weekday())
        pazar = pazartesi + timedelta(days=6)
        # Iki ay kapsiyorsa her iki ayin kisaltmasini goster
        if pazartesi.month == pazar.month:
            tarih_str = f"{pazartesi.day}-{pazar.day} {AY_KISA.get(pazartesi.month, '')}"
        else:
            tarih_str = f"{pazartesi.day} {AY_KISA.get(pazartesi.month, '')} - {pazar.day} {AY_KISA.get(pazar.month, '')}"
        label = f"{iso_hafta}. Hafta ({tarih_str})"
        # iso_yil kullan (takvim yili degil!) - yil sinirlarinda kritik
        result.append({'hafta': iso_hafta, 'iso_yil': iso_yil, 'label': label})
    return result


def _get_bilanco(yil, hafta):
    with get_db() as conn:
        row = conn.execute(
            'SELECT * FROM haftalik_bilanco WHERE yil=%s AND hafta=%s', (yil, hafta)
        ).fetchone()
        if row:
            return dict(row)
        return None


def _get_kalemler(bilanco_id):
    with get_db() as conn:
        rows = conn.execute(
            'SELECT * FROM haftalik_bilanco_kalem WHERE bilanco_id=%s ORDER BY id', (bilanco_id,)
        ).fetchall()
        return [dict(r) for r in rows]


def _save_bilanco(yil, hafta, papel_fiyat, tutkal_fiyat, kalemler):
    with get_db() as conn:
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row = conn.execute(
            'SELECT id FROM haftalik_bilanco WHERE yil=%s AND hafta=%s', (yil, hafta)
        ).fetchone()
        if row:
            bilanco_id = row['id']
            conn.execute(
                'UPDATE haftalik_bilanco SET papel_fiyat=%s, tutkal_fiyat=%s, updated_at=%s WHERE id=%s',
                (papel_fiyat, tutkal_fiyat, now, bilanco_id)
            )
            conn.execute('DELETE FROM haftalik_bilanco_kalem WHERE bilanco_id=%s', (bilanco_id,))
        else:
            cur = conn.execute(
                'INSERT INTO haftalik_bilanco (yil, hafta, papel_fiyat, tutkal_fiyat, created_at, updated_at) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id',
                (yil, hafta, papel_fiyat, tutkal_fiyat, now, now)
            )
            bilanco_id = cur.fetchone()['id']

        for k in kalemler:
            if float(k.get('adet', 0) or 0) <= 0:
                continue
            desi = float(k.get('desi_degeri', 0) or 0)
            adet = float(k.get('adet', 0) or 0)
            satis_fiyat = float(k.get('satis_fiyat', 0) or 0)
            tutkal = float(k.get('tutkal', 0) or tutkal_fiyat)
            papel_ham = desi * papel_fiyat
            ham_maliyet = papel_ham + tutkal
            fark = (satis_fiyat - ham_maliyet) if satis_fiyat > 0 else 0
            kazanc = fark * adet
            haftalik_hammadde = desi * adet
            haftalik_satis = satis_fiyat * adet

            conn.execute('''
                INSERT INTO haftalik_bilanco_kalem
                (bilanco_id, urun_kod, urun_ad, desi, adet, papel_fiyat, tutkal, satis_fiyat,
                 papel_ham_fiyat, ham_maliyet, fark, kazanc, haftalik_hammadde, haftalik_satis)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ''', (bilanco_id, k['kod'], k['ad'], desi, adet, papel_fiyat, tutkal, satis_fiyat,
                  papel_ham, ham_maliyet, fark, kazanc, haftalik_hammadde, haftalik_satis))

        return bilanco_id


def _build_excel(baslik, desi_urunler, inputs, inp_papel, inp_tutkal):
    """openpyxl ile Excel dosyasi olustur, bytes dondur."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = 'Haftalık Bilanço'

    ws.merge_cells('A1:K1')
    ws['A1'] = baslik
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    pf = float(inp_papel.value or 25)
    tf = float(inp_tutkal.value or 10)
    ws['A2'] = f'Papel Fiyat: {pf:.2f} TL/desi'
    ws['D2'] = f'Tutkal: {tf:.0f} TL'
    ws['A2'].font = Font(bold=True, color='FF6600')
    ws['D2'].font = Font(bold=True, color='FF6600')

    headers = ['Ürün', 'DESİ', 'Adet', 'Papel Ham', 'Tutkal', 'Ham Maliyet', 'Satış Fiyat', 'Kar', 'Kazanç', 'Hammadde (desi)', 'Ciro']
    header_fill = PatternFill(start_color='1C4461', end_color='1C4461', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    toplam_hammadde = 0
    toplam_ciro = 0
    toplam_kazanc = 0
    for row_idx, u in enumerate(desi_urunler, 5):
        kod = u['kod']
        desi = float(u.get('desi_degeri', 0) or 0)
        adet = float(inputs[f"{kod}_adet"].value or 0)
        sf = float(inputs[f"{kod}_satis"].value or 0)
        tk = float(inputs[f"{kod}_tutkal"].value or 0)
        ph = desi * pf
        hm = ph + tk
        kar = (sf - hm) if sf > 0 else 0  # FIX: UI ile tutarli
        kazanc = kar * adet
        hammadde = desi * adet
        ciro = sf * adet
        toplam_hammadde += hammadde
        toplam_ciro += ciro
        toplam_kazanc += kazanc

        values = [u['ad'], desi, adet, ph, tk, hm, sf, kar, kazanc, hammadde, ciro]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx >= 2:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            if col_idx == 8:
                cell.font = Font(bold=True, color='2E7D32' if kar >= 0 else 'C10015')

    t_row = 5 + len(desi_urunler)
    ws.cell(row=t_row, column=1, value='TOPLAM').font = Font(bold=True, size=11)
    ws.cell(row=t_row, column=9, value=toplam_kazanc).font = Font(bold=True, size=11)
    ws.cell(row=t_row, column=9).number_format = '#,##0.00'
    ws.cell(row=t_row, column=10, value=toplam_hammadde).font = Font(bold=True)
    ws.cell(row=t_row, column=10).number_format = '#,##0.00'
    ws.cell(row=t_row, column=11, value=toplam_ciro).font = Font(bold=True, size=11)
    ws.cell(row=t_row, column=11).number_format = '#,##0.00'

    ws.cell(row=t_row + 1, column=1, value=f'Toplam Hammadde: {toplam_hammadde:.1f} desi = {toplam_hammadde/1000:.3f} m³')
    ws.cell(row=t_row + 1, column=1).font = Font(bold=True, color='1565C0')

    widths = [20, 8, 8, 12, 8, 12, 12, 10, 12, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@ui.page('/haftalik-bilanco')
def haftalik_bilanco_page():
    if not create_layout(active_path='/haftalik-bilanco', page_title='Haftalık Bilanço'):
        return

    ayarlar = get_company_settings()
    if not ayarlar.get('uretim_takibi'):
        with ui.column().classes('w-full q-pa-md'):
            ui.label('Üretim takibi bu firma için aktif değil.').classes('text-h6 text-grey-7')
            ui.label('Ayarlar sayfasından "Üretim Takibi" seçeneğini açın.').classes('text-caption')
        return

    now = datetime.now()
    # state['yil'] artik ISO yil - DB key olarak iso_yil kullanilir
    state = {'yil': now.year, 'ay': now.month, 'hafta': 0, 'iso_yil': now.year}
    table_container = None

    # Papel/tutkal varsayilan degerler
    DEFAULT_PAPEL = 25
    DEFAULT_TUTKAL = 10

    with ui.column().classes('w-full q-pa-sm').style('max-width: 1400px; margin: 0 auto'):
        # Baslik + donem secici
        with ui.card().classes('w-full q-pa-sm q-mb-sm'):
            with ui.row().classes('w-full items-center gap-2 flex-wrap'):
                ui.label('Haftalık Bilanço').classes('text-h6 text-weight-bold')
                ui.space()

                yil_options = {y: str(y) for y in range(now.year - 2, now.year + 2)}
                ay_options = {m: AY_ISIMLERI[m] for m in range(1, 13)}

                sel_ay = ui.select(options=ay_options, value=state['ay'], label='Ay').props('outlined dense').classes('w-32')
                sel_yil = ui.select(options=yil_options, value=state['yil'], label='Yıl').props('outlined dense').classes('w-28')

                haftalar = _get_haftalar(state['yil'], state['ay'])
                bugun_iso = date.today().isocalendar()
                bugun_hafta = bugun_iso[1]
                default_h = haftalar[0]['hafta'] if haftalar else 1
                default_iso_yil = haftalar[0]['iso_yil'] if haftalar else now.year
                for h in haftalar:
                    if h['hafta'] == bugun_hafta:
                        default_h = h['hafta']
                        default_iso_yil = h['iso_yil']
                        break
                state['hafta'] = default_h
                state['iso_yil'] = default_iso_yil
                # Hafta secici: value olarak "iso_yil:hafta" formatinda unique key kullan
                hafta_opts = {f"{h['iso_yil']}:{h['hafta']}": h['label'] for h in haftalar}
                default_key = f"{default_iso_yil}:{default_h}"
                sel_hafta = ui.select(options=hafta_opts, value=default_key, label='Hafta').props('outlined dense').classes('w-56')

        # Papel fiyat + Tutkal girisi
        with ui.card().classes('w-full q-pa-sm q-mb-sm').style('background: #FFF8E1; border: 1px solid #FFE082'):
            with ui.row().classes('w-full items-center gap-md flex-wrap'):
                ui.icon('price_change', color='orange-8').style('font-size: 24px')
                ui.label('Hammadde Fiyatları').classes('text-subtitle2 text-weight-bold text-orange-9')
                inp_papel = ui.number('Papel Fiyat (TL/desi)', value=DEFAULT_PAPEL, format='%.2f').props('outlined dense').classes('w-48')
                inp_tutkal = ui.number('Tutkal (TL/birim)', value=DEFAULT_TUTKAL, format='%.2f').props('outlined dense').classes('w-48')

        # Tablo
        table_container = ui.column().classes('w-full')

        # Event listener referanslari - birikmesini onlemek icin
        _papel_handler = {'ref': None}
        _tutkal_handler = {'ref': None}

        def load_bilanco():
            table_container.clear()

            # Onceki event listener'lari kaldir (birikmesini onle)
            if _papel_handler['ref']:
                try:
                    inp_papel.on_value_change(_papel_handler['ref'], remove=True)
                except Exception:
                    pass
            if _tutkal_handler['ref']:
                try:
                    inp_tutkal.on_value_change(_tutkal_handler['ref'], remove=True)
                except Exception:
                    pass

            urunler = get_urun_list()
            desi_urunler = [u for u in urunler if float(u.get('desi_degeri', 0) or 0) > 0]

            # DB'den iso_yil ile sorgula (takvim yili degil!)
            bilanco = _get_bilanco(state['iso_yil'], state['hafta'])
            kalemler = _get_kalemler(bilanco['id']) if bilanco else []
            kalem_map = {k['urun_kod']: k for k in kalemler}

            if bilanco:
                inp_papel.value = float(bilanco.get('papel_fiyat', DEFAULT_PAPEL) or DEFAULT_PAPEL)
                inp_tutkal.value = float(bilanco.get('tutkal_fiyat', DEFAULT_TUTKAL) or DEFAULT_TUTKAL)
            else:
                # Kayitli bilanco yoksa varsayilana don (onceki haftanin degerlerini tasima)
                inp_papel.value = DEFAULT_PAPEL
                inp_tutkal.value = DEFAULT_TUTKAL

            rows_data = []
            for u in desi_urunler:
                k = kalem_map.get(u['kod'], {})
                rows_data.append({
                    'kod': u['kod'],
                    'ad': u['ad'],
                    'desi_degeri': float(u.get('desi_degeri', 0) or 0),
                    'adet': float(k.get('adet', 0) or 0),
                    'satis_fiyat': float(k.get('satis_fiyat', 0) or 0),
                    'tutkal': float(k.get('tutkal', 0) or inp_tutkal.value or DEFAULT_TUTKAL),
                })

            with table_container:
                if not desi_urunler:
                    ui.label('DESİ değeri olan ürün bulunamadı. Stok sayfasından ürünlere DESİ değeri girin.').classes('text-grey-7 q-pa-md')
                    return

                inputs = {}
                row_labels = {}
                ozet_labels = {}

                def _recalc_row(kod, desi):
                    pf = float(inp_papel.value or DEFAULT_PAPEL)
                    adet = float(inputs[f"{kod}_adet"].value or 0)
                    sf = float(inputs[f"{kod}_satis"].value or 0)
                    tk = float(inputs[f"{kod}_tutkal"].value or 0)
                    ph = desi * pf
                    hm = ph + tk
                    kar = (sf - hm) if sf > 0 else 0
                    kazanc = kar * adet
                    hammadde = desi * adet
                    ciro = sf * adet
                    lbls = row_labels.get(kod)
                    if lbls:
                        lbls['ph'].set_text(fmt_para(ph))
                        lbls['hm'].set_text(fmt_para(hm))
                        lbls['kar'].set_text(fmt_para(kar))
                        lbls['kazanc'].set_text(fmt_para(kazanc))
                        lbls['hammadde'].set_text(fmt_para(hammadde))
                        lbls['ciro'].set_text(fmt_para(ciro))
                    _recalc_totals()

                def _recalc_all(_=None):
                    for r in rows_data:
                        _recalc_row(r['kod'], r['desi_degeri'])

                def _recalc_totals():
                    t_hammadde = 0
                    t_ciro = 0
                    t_kazanc = 0
                    pf = float(inp_papel.value or DEFAULT_PAPEL)
                    for r in rows_data:
                        kod = r['kod']
                        desi = r['desi_degeri']
                        adet = float(inputs[f"{kod}_adet"].value or 0)
                        sf = float(inputs[f"{kod}_satis"].value or 0)
                        tk = float(inputs[f"{kod}_tutkal"].value or 0)
                        hm = desi * pf + tk
                        kar = (sf - hm) if sf > 0 else 0
                        t_hammadde += desi * adet
                        t_ciro += sf * adet
                        t_kazanc += kar * adet
                    if ozet_labels:
                        ozet_labels['hammadde'].set_text(f'Toplam Hammadde: {fmt_para(t_hammadde)} desi = {fmt_para(t_hammadde/1000)} m³')
                        ozet_labels['ciro'].set_text(f'Toplam Ciro: {fmt_para(t_ciro)} TL')
                        ozet_labels['kazanc'].set_text(f'Toplam Kazanç: {fmt_para(t_kazanc)} TL')

                COL_W = ['14%', '6%', '8%', '10%', '8%', '10%', '10%', '9%', '10%', '8%', '7%']
                COL_NAMES = ['Ürün', 'DESİ', 'Adet', 'Papel Ham', 'Tutkal', 'Ham Maliyet', 'Satış Fiyat', 'Kar', 'Kazanç', 'Hammadde', 'Ciro']

                with ui.card().classes('w-full q-pa-none'):
                    with ui.element('div').classes('w-full').style('overflow-x: auto'):
                        with ui.element('table').style('width:100%; border-collapse:collapse; font-size:13px; table-layout:fixed'):
                            with ui.element('colgroup'):
                                for w in COL_W:
                                    ui.element('col').style(f'width:{w}')

                            with ui.element('thead'):
                                with ui.element('tr').style('background: linear-gradient(180deg, #1c4461 0%, #1a3a52 100%); color:white'):
                                    for col in COL_NAMES:
                                        align = 'right' if col != 'Ürün' else 'left'
                                        with ui.element('th').style(f'padding:8px 6px; text-align:{align}; font-weight:600; white-space:nowrap; font-size:11px'):
                                            ui.label(col)

                            with ui.element('tbody'):
                                for idx, r in enumerate(rows_data):
                                    bg = '#f8f9fa' if idx % 2 == 0 else '#ffffff'
                                    kod = r['kod']
                                    desi = r['desi_degeri']

                                    with ui.element('tr').style(f'background:{bg}; border-bottom:1px solid #e0e0e0'):
                                        with ui.element('td').style('padding:6px 6px; font-weight:500; overflow:hidden; text-overflow:ellipsis; white-space:nowrap'):
                                            ui.label(r['ad'])
                                        with ui.element('td').style('padding:6px 6px; text-align:right; color:#1565C0; font-weight:600'):
                                            ui.label(f"{desi:.1f}")
                                        with ui.element('td').style('padding:2px 4px; text-align:right'):
                                            inp_a = ui.number(value=r['adet'], format='%.0f').props('dense borderless input-style="text-align:right"').style('width:100%')
                                            inputs[f"{kod}_adet"] = inp_a
                                        with ui.element('td').style('padding:6px 6px; text-align:right'):
                                            lbl_ph = ui.label(fmt_para(desi * float(inp_papel.value or DEFAULT_PAPEL))).classes('text-caption')
                                        with ui.element('td').style('padding:2px 4px; text-align:right'):
                                            inp_t = ui.number(value=r['tutkal'], format='%.0f').props('dense borderless input-style="text-align:right"').style('width:100%')
                                            inputs[f"{kod}_tutkal"] = inp_t
                                        with ui.element('td').style('padding:6px 6px; text-align:right'):
                                            hm = desi * float(inp_papel.value or DEFAULT_PAPEL) + r['tutkal']
                                            lbl_hm = ui.label(fmt_para(hm)).classes('text-caption')
                                        with ui.element('td').style('padding:2px 4px; text-align:right'):
                                            inp_sf = ui.number(value=r['satis_fiyat'], format='%.0f').props('dense borderless input-style="text-align:right"').style('width:100%')
                                            inputs[f"{kod}_satis"] = inp_sf
                                        with ui.element('td').style('padding:6px 6px; text-align:right; color:#2E7D32; font-weight:600'):
                                            kar = (r['satis_fiyat'] - hm) if r['satis_fiyat'] > 0 else 0
                                            lbl_kar = ui.label(fmt_para(kar)).classes('text-caption')
                                        with ui.element('td').style('padding:6px 6px; text-align:right; color:#1565C0; font-weight:600'):
                                            lbl_kazanc = ui.label(fmt_para(kar * r['adet'])).classes('text-caption')
                                        with ui.element('td').style('padding:6px 6px; text-align:right'):
                                            lbl_hammadde = ui.label(fmt_para(desi * r['adet'])).classes('text-caption')
                                        with ui.element('td').style('padding:6px 6px; text-align:right; font-weight:600'):
                                            lbl_ciro = ui.label(fmt_para(r['satis_fiyat'] * r['adet'])).classes('text-caption')

                                    row_labels[kod] = {
                                        'ph': lbl_ph, 'hm': lbl_hm, 'kar': lbl_kar,
                                        'kazanc': lbl_kazanc, 'hammadde': lbl_hammadde, 'ciro': lbl_ciro,
                                    }

                                    inp_a.on_value_change(lambda _, k=kod, d=desi: _recalc_row(k, d))
                                    inp_t.on_value_change(lambda _, k=kod, d=desi: _recalc_row(k, d))
                                    inp_sf.on_value_change(lambda _, k=kod, d=desi: _recalc_row(k, d))

                # Ozet
                toplam_hammadde = sum(r['desi_degeri'] * r['adet'] for r in rows_data)
                toplam_ciro = sum(r['satis_fiyat'] * r['adet'] for r in rows_data)
                toplam_kazanc = sum(
                    ((r['satis_fiyat'] - (r['desi_degeri'] * float(inp_papel.value or DEFAULT_PAPEL) + r['tutkal'])) if r['satis_fiyat'] > 0 else 0) * r['adet']
                    for r in rows_data
                )
                with ui.card().classes('w-full q-pa-sm q-mt-sm').style('background:#E3F2FD'):
                    with ui.row().classes('w-full items-center gap-4 flex-wrap'):
                        ozet_labels['hammadde'] = ui.label(
                            f'Toplam Hammadde: {fmt_para(toplam_hammadde)} desi = {fmt_para(toplam_hammadde/1000)} m³'
                        ).classes('text-subtitle2 text-weight-bold')
                        ozet_labels['ciro'] = ui.label(
                            f'Toplam Ciro: {fmt_para(toplam_ciro)} TL'
                        ).classes('text-subtitle2 text-weight-bold text-green-8')
                        ozet_labels['kazanc'] = ui.label(
                            f'Toplam Kazanç: {fmt_para(toplam_kazanc)} TL'
                        ).classes('text-subtitle2 text-weight-bold text-blue-8')

                # Papel/tutkal on_change - her load_bilanco'da yeni handler atanir
                _papel_handler['ref'] = _recalc_all
                _tutkal_handler['ref'] = _recalc_all
                inp_papel.on_value_change(_recalc_all)
                inp_tutkal.on_value_change(_recalc_all)

                # Butonlar
                with ui.row().classes('w-full justify-end q-mt-sm gap-sm'):
                    def _kaydet():
                        try:
                            kalem_list = []
                            for u in desi_urunler:
                                kod = u['kod']
                                kalem_list.append({
                                    'kod': kod,
                                    'ad': u['ad'],
                                    'desi_degeri': float(u.get('desi_degeri', 0) or 0),
                                    'adet': float(inputs[f"{kod}_adet"].value or 0),
                                    'satis_fiyat': float(inputs[f"{kod}_satis"].value or 0),
                                    'tutkal': float(inputs[f"{kod}_tutkal"].value or 0),
                                })
                            # iso_yil kullan (takvim yili degil!)
                            _save_bilanco(
                                state['iso_yil'], state['hafta'],
                                float(inp_papel.value or DEFAULT_PAPEL),
                                float(inp_tutkal.value or DEFAULT_TUTKAL),
                                kalem_list
                            )
                            notify_ok('Bilanço kaydedildi')
                            load_bilanco()
                        except Exception as e:
                            notify_err(f'Hata: {e}')

                    def _pdf():
                        try:
                            from services.pdf_service import generate_table_pdf, save_pdf_preview
                            headers = ['Ürün', 'DESİ', 'Adet', 'Papel Ham', 'Tutkal', 'Ham Mal.', 'Satış F.', 'Kar', 'Kazanç', 'Hammadde', 'Ciro']
                            pdf_rows = []
                            for u in desi_urunler:
                                kod = u['kod']
                                desi = float(u.get('desi_degeri', 0) or 0)
                                adet = float(inputs[f"{kod}_adet"].value or 0)
                                sf = float(inputs[f"{kod}_satis"].value or 0)
                                tk = float(inputs[f"{kod}_tutkal"].value or 0)
                                pf = float(inp_papel.value or DEFAULT_PAPEL)
                                ph = desi * pf
                                hm = ph + tk
                                kar = (sf - hm) if sf > 0 else 0  # FIX: UI ile tutarli
                                pdf_rows.append([
                                    u['ad'], f'{desi:.1f}', f'{adet:.0f}',
                                    f'{ph:.2f}', f'{tk:.0f}', f'{hm:.2f}',
                                    f'{sf:.0f}', f'{kar:.2f}', f'{kar*adet:.2f}',
                                    f'{desi*adet:.1f}', f'{sf*adet:.0f}'
                                ])
                            baslik = f"Haftalık Bilanço - {state['hafta']}. Hafta {state['iso_yil']}"
                            pdf_bytes = generate_table_pdf(baslik, headers, pdf_rows)
                            preview_url = save_pdf_preview(pdf_bytes, f'bilanco_{state["iso_yil"]}_{state["hafta"]}.pdf')
                            ui.run_javascript(f"window.open('{preview_url}', '_blank')")
                        except Exception as e:
                            notify_err(f'PDF hatası: {e}')

                    def _excel():
                        try:
                            from services.pdf_service import get_pdf_preview_dir
                            baslik = f"Haftalık Bilanço - {state['hafta']}. Hafta {state['iso_yil']}"
                            excel_bytes = _build_excel(baslik, desi_urunler, inputs, inp_papel, inp_tutkal)
                            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                            fname = f'bilanco_{state["iso_yil"]}_{state["hafta"]}_{ts}.xlsx'
                            out_dir = get_pdf_preview_dir()
                            out_path = out_dir / fname
                            out_path.write_bytes(excel_bytes)
                            ui.run_javascript(f"window.open('/pdf-preview/{fname}', '_blank')")
                        except Exception as e:
                            notify_err(f'Excel hatası: {e}')

                    ui.button('Kaydet', icon='save', color='primary', on_click=_kaydet).props('unelevated')
                    ui.button('PDF', icon='picture_as_pdf', color='secondary', on_click=_pdf).props('dense')
                    ui.button('Excel', icon='table_chart', color='green-7', on_click=_excel).props('dense')

        # Donem degisince
        def on_donem():
            state['yil'] = sel_yil.value
            state['ay'] = sel_ay.value
            yeni_haftalar = _get_haftalar(state['yil'], state['ay'])
            new_opts = {f"{h['iso_yil']}:{h['hafta']}": h['label'] for h in yeni_haftalar}
            sel_hafta.options = new_opts
            if yeni_haftalar:
                first_key = f"{yeni_haftalar[0]['iso_yil']}:{yeni_haftalar[0]['hafta']}"
                sel_hafta.value = first_key
            sel_hafta.update()
            _parse_hafta_key(sel_hafta.value)
            load_bilanco()

        def _parse_hafta_key(key):
            """'iso_yil:hafta' formatindaki key'i parse et."""
            if not key or ':' not in str(key):
                return
            parts = str(key).split(':')
            state['iso_yil'] = int(parts[0])
            state['hafta'] = int(parts[1])

        def on_hafta():
            _parse_hafta_key(sel_hafta.value)
            load_bilanco()

        sel_ay.on_value_change(lambda _: on_donem())
        sel_yil.on_value_change(lambda _: on_donem())
        sel_hafta.on_value_change(lambda _: on_hafta())

        load_bilanco()
