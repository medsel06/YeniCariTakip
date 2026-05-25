# -*- coding: utf-8 -*-
"""Odeme Plani'ni cari'ye bagli yeniden isle ( on muhasebe duzeltmesi).

- Mevcut kategori='ÖDEME PLANI' gelir_gider (+ bagli kasa) kayitlarini siler.
- Odeme Plani sheet'ini tekrar isler:
    * Gercek tedarikci/musteri -> cari kart (yoksa ac) + vadeli gelir_gider (firma_kod dolu)
    * Saf kredi karti harcamalari (YKB/Bonuscard/market/cayci vb) -> ATLA (KK Harcamasi'nda zaten var)
"""
import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from datetime import datetime
import psycopg2
from db import DB_CONFIG
from openpyxl import load_workbook

COMMIT = '--commit' in sys.argv
PATH = os.environ.get('XLSX_PATH', r"C:\Users\ykahm\OneDrive\Masaüstü\Kasa defteri.xlsx")
NOW = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def _norm(s):
    s = (s or '').lower()
    for a, b in [('ü','u'),('ö','o'),('ş','s'),('ı','i'),('ç','c'),('ğ','g'),('İ','i')]:
        s = s.replace(a, b)
    return s


# Gercek cari eslestirme: keyword -> (kod, ad). MKE mevcut F006.
CARI_ESLEME = [
    ('aybimas', ('CARI-AYBIMAS', 'Aybimaş AŞ')),
    ('mke',     ('F006', 'MKE AŞ')),
    ('dilek',   ('CARI-DILEK', 'Dilek Geri Dönüşüm')),
    ('oknal',   ('CARI-OKNAL', 'Oknal Gaz')),
    ('ismail altiparmak', ('CARI-ISMAIL', 'İsmail Altıparmak')),
]

# Saf kredi karti / harcama isaretleri (KK Harcamasi'nda var -> atla)
KART_ISARET = ['ykb', 'bonus', 'kuveyt', 'qnb', 'akbank', 'ziraat kart', 'market harcama',
               'cayci', 'somun', 'ayyildiz', 'sanayi tupu', 'yemek ucret', 'spiral']


def cari_bul(aciklama):
    a = _norm(aciklama)
    for key, (kod, ad) in CARI_ESLEME:
        if key in a:
            return kod, ad
    return None, None


def kart_mi(aciklama):
    a = _norm(aciklama)
    return any(k in a for k in KART_ISARET)


def num(v):
    try:
        return float(str(v).replace(',', '.')) if v not in (None, '') else 0.0
    except ValueError:
        return 0.0


c = psycopg2.connect(**DB_CONFIG); c.autocommit = False; cur = c.cursor()
cur.execute("SET search_path TO t_6, public")

# 1) Eski ÖDEME PLANI temizle
cur.execute("SELECT id FROM gelir_gider WHERE kategori='ÖDEME PLANI'")
eski_ids = [r[0] for r in cur.fetchall()]
print(f"Silinecek eski ÖDEME PLANI gelir_gider: {len(eski_ids)} (+ bagli kasa)")
if COMMIT and eski_ids:
    cur.execute("DELETE FROM kasa WHERE gelir_gider_id = ANY(%s)", (eski_ids,))
    cur.execute("DELETE FROM gelir_gider WHERE kategori='ÖDEME PLANI'")

# 2) Sheet'i yeniden isle
wb = load_workbook(PATH, data_only=True)
ws = wb['Ödeme Planı']
cari_olustur = {}
eklenecek = []
atlanan = []
for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    kisi = str(r[1] or '').strip()
    tutar = num(r[3])
    if not kisi or tutar <= 0:
        continue
    tarih = str(r[0]).split()[0] if r[0] else ''
    ba = _norm(str(r[2] or ''))
    odendi = (str(r[5] or '').strip().upper() == 'ÖDENDİ') if len(r) > 5 else False
    kod, ad = cari_bul(kisi)
    if not kod:
        if kart_mi(kisi):
            atlanan.append((kisi, tutar, 'kredi karti -> KK Harcamasi\'nda var'))
        else:
            # tanimsiz gercek cari: kisi adiyla cari ac
            kod = 'CARI-' + _norm(kisi)[:20].replace(' ', '-')
            ad = kisi
    if not kod:
        continue
    if kart_mi(kisi) and not cari_bul(kisi)[0]:
        continue  # kart kalemi atla
    tur = 'GELIR' if 'alacak' in ba else 'GIDER'
    if kod not in ('F006',):
        cari_olustur[kod] = ad
    eklenecek.append((tarih, tur, kisi, tutar, 'ODENDI' if odendi else 'ODENMEDI',
                      '' if odendi else tarih, kod, ad))

print(f"\nCARI bazli eklenecek: {len(eklenecek)}")
toplam_b = sum(e[3] for e in eklenecek if e[1] == 'GIDER')
toplam_a = sum(e[3] for e in eklenecek if e[1] == 'GELIR')
print(f"  Borç={toplam_b:,.2f}  Alacak={toplam_a:,.2f}")
print(f"\nAcilacak cari kartlar: {cari_olustur}")
print(f"\nATLANAN (kredi karti, KK'da var): {len(atlanan)}")
for k, t, neden in atlanan:
    print(f"  - {t:>12,.2f} | {k[:45]} ({neden})")

# cari bazli ozet
from collections import defaultdict
ozet = defaultdict(float)
for e in eklenecek:
    ozet[e[7]] += e[3] if e[1] == 'GIDER' else -e[3]
print(f"\nCARI BAZLI NET (borç+/alacak-):")
for ad, net in sorted(ozet.items(), key=lambda x: -abs(x[1])):
    print(f"  {ad}: {net:,.2f}")

if COMMIT:
    # cari kartlari ac
    for kod, ad in cari_olustur.items():
        cur.execute("SELECT 1 FROM firmalar WHERE kod=%s", (kod,))
        if not cur.fetchone():
            cur.execute("INSERT INTO firmalar (kod, ad, aktif) VALUES (%s,%s,1)", (kod, ad))
    # gelir_gider ekle (vadeli, cari bagli). ODENDI ise bagli kasa.
    for tarih, tur, kisi, tutar, durum, vade, kod, ad in eklenecek:
        cur.execute(
            "INSERT INTO gelir_gider (tarih, tur, kategori, aciklama, tutar, toplam, odeme_sekli, "
            "firma_kod, firma_ad, odeme_durumu, vade_tarih, created_at) "
            "VALUES (%s,%s,'ÖDEME PLANI',%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (tarih, tur, kisi, tutar, tutar, 'NAKIT' if durum == 'ODENDI' else '',
             kod, ad, durum, vade, NOW))
        gg_id = cur.fetchone()[0]
        if durum == 'ODENDI':
            kt = 'GIDER' if tur == 'GIDER' else 'GELIR'
            cur.execute(
                "INSERT INTO kasa (tarih, firma_kod, firma_ad, tur, tutar, odeme_sekli, aciklama, gelir_gider_id, created_at) "
                "VALUES (%s,%s,%s,%s,%s,'NAKIT',%s,%s,%s)",
                (tarih, kod, ad, kt, tutar, f"Ödeme planı: {kisi}"[:200], gg_id, NOW))
    c.commit()
    print("\n✅ COMMIT edildi — ÖDEME PLANI cari'ye baglandi.")
else:
    c.rollback()
    print("\n(DRY-RUN — --commit ile uygula)")
c.close()
