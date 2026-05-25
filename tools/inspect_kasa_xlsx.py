# -*- coding: utf-8 -*-
"""Kasa defteri.xlsx incele — sheet adlari ve sutun yapisi."""
import os
import sys

from openpyxl import load_workbook

PATH = r"C:\Users\ykahm\OneDrive\Masaüstü\Kasa defteri.xlsx"

if not os.path.exists(PATH):
    print(f"FILE NOT FOUND: {PATH}")
    sys.exit(1)

wb = load_workbook(PATH, data_only=True)
print(f"SHEETS ({len(wb.sheetnames)}):")
for sn in wb.sheetnames:
    print(f"  - {sn!r}")
print()


def trunc(c, n=50):
    if c is None:
        return ""
    s = str(c)
    return s if len(s) <= n else s[:n - 3] + "..."


for sn in wb.sheetnames:
    ws = wb[sn]
    print("=" * 80)
    print(f"SHEET: {sn!r}")
    print(f"  dimensions={ws.dimensions} max_row={ws.max_row} max_col={ws.max_column}")
    print()
    # ilk 10 satir
    n_preview = min(12, ws.max_row)
    print(f"  --- ilk {n_preview} satir ---")
    for i, r in enumerate(ws.iter_rows(min_row=1, max_row=n_preview, values_only=True), 1):
        cells = [trunc(c) for c in r]
        print(f"  R{i:>3}: {cells}")
    print()
    # Doluluk: kolon basina dolu hucre sayisi
    fill = [0] * ws.max_column
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        for ci, c in enumerate(r):
            if c is not None and str(c).strip():
                fill[ci] += 1
    print(f"  --- kolon doluluk sayilari ---")
    print(f"  {fill}")
    print()
    # Son 3 satir
    if ws.max_row > n_preview:
        print(f"  --- son 3 satir ---")
        start = max(n_preview + 1, ws.max_row - 2)
        for i, r in enumerate(ws.iter_rows(min_row=start, max_row=ws.max_row, values_only=True), start):
            cells = [trunc(c) for c in r]
            print(f"  R{i:>3}: {cells}")
        print()
