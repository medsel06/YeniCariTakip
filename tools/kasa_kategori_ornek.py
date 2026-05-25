# -*- coding: utf-8 -*-
"""Belirsiz kategorilerin ornek satirlarini goster (parse mantigi icin)."""
from openpyxl import load_workbook

PATH = r"C:\Users\ykahm\OneDrive\Masaüstü\Kasa defteri.xlsx"
wb = load_workbook(PATH, data_only=True)


def dump(sheet, kategoriler, kol_turu=5):
    ws = wb[sheet]
    print(f"\n{'='*70}\nSHEET: {sheet}")
    for i, r in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
        turu = str(r[kol_turu] or '').strip()
        if turu in kategoriler:
            tarih = str(r[0]).split()[0] if r[0] else ''
            giren = r[2] or ''
            cikan = r[3] or ''
            kg = r[7] if len(r) > 7 else ''
            print(f"  [{turu}] {tarih} gir={giren} cik={cikan} kg={kg} | {str(r[1] or '')[:55]}")


# kasa: Banka, Maas, KK Harcamasi, Mal Alimi, Ortaktan Alinan, Ortaklara Odenen, Murat eski borc
dump('kasa', {'Banka', 'Mal Alımı', 'KK Harcaması'})
print("\n--- MAAS (ilk 12) ---")
ws = wb['kasa']
c = 0
for i, r in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    if str(r[5] or '').strip() == 'Maaş':
        tarih = str(r[0]).split()[0] if r[0] else ''
        print(f"  {tarih} cik={r[3]} | {str(r[1] or '')[:55]}")
        c += 1
        if c >= 12:
            break
print("\n--- ORTAKTAN ALINAN / ORTAKLARA ODENEN / MURAT ESKI BORC ---")
for i, r in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    t = str(r[5] or '').strip()
    if t in ('Ortaktan Alınan', 'Ortaklara Ödenen', 'Murat eski borç'):
        tarih = str(r[0]).split()[0] if r[0] else ''
        print(f"  [{t}] {tarih} gir={r[2]} cik={r[3]} | {str(r[1] or '')[:50]}")

# banka sheet Turu kol=5
print("\n\n### BANKA SHEET Turu dagilimi ###")
from collections import Counter
wb2 = wb['banka']
cnt = Counter()
for r in wb2.iter_rows(min_row=2, max_row=wb2.max_row, values_only=True):
    t = str(r[5] or '').strip()
    if t:
        cnt[t] += 1
for k, v in cnt.most_common():
    print(f"  {v:>4} {k}")
