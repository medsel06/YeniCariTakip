# -*- coding: utf-8 -*-
"""kasa sheet'inde Turu kolonu unique listesi + Satis satirlarini listele."""
from openpyxl import load_workbook
from collections import Counter

PATH = r"C:\Users\ykahm\OneDrive\Masaüstü\Kasa defteri.xlsx"

wb = load_workbook(PATH, data_only=True)
ws = wb["kasa"]

print(f"max_row={ws.max_row}")
print()

# Tum Turu degerleri (kolon F = index 5)
turu_counter = Counter()
satislar = []
for i, r in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
    turu = r[5]
    if turu is None or not str(turu).strip():
        continue
    turu_s = str(turu).strip()
    turu_counter[turu_s] += 1
    if turu_s == "Satış" or turu_s.lower() == "satis" or "sat" in turu_s.lower():
        satislar.append((i, r))

print("=== Turu unique degerleri ===")
for k, v in sorted(turu_counter.items(), key=lambda x: -x[1]):
    print(f"  {v:>4}  {k!r}")
print()

print(f"=== 'Satis' iceren satirlar ({len(satislar)}) ===")
for i, r in satislar:
    tarih = str(r[0]).split()[0] if r[0] else ""
    aciklama = str(r[1] or "")[:50]
    giren = r[2] or ""
    cikan = r[3] or ""
    turu = r[5] or ""
    odemetar = str(r[6]).split()[0] if r[6] else ""
    kg = r[7] or ""
    print(f"  R{i:>3} {tarih:>10} | gir={giren} cik={cikan} | kg={kg} | tur={turu!r} | odetar={odemetar} | {aciklama!r}")
