# -*- coding: utf-8 -*-
"""t_6 (MURAT BAYSAL) icin 'Kasa defteri.xlsx' migration.

KULLANIM:
    python tools/migrate_excel_t6.py            # DRY-RUN (rapor, DB'ye yazmaz)
    python tools/migrate_excel_t6.py --commit   # GERCEK yazar (once test verisini temizler)

KATEGORI KURALLARI (kullanici onayli):
  Satis          -> hareketler(SATIS) MOTOSIKLET PARCASI + kasa(GELIR)
  Masraf         -> gelir_gider(GIDER) + kasa(GIDER)
  Maas (isci)    -> personel + personel_hareket + gelir_gider(GIDER, ISCILIK) [gider olarak da]
  Banka          -> transfer (nakit<->BANKA): bankaya yatan=CIKAN nakit->BANKA, cekilen=GIREN BANKA->nakit
  KK Harcamasi   -> kredi karti hesabi (Ziraat/Akbank/QNB/Kuveytturk), tutar aciklamadan
  Ortaktan Alinan-> kasa(GELIR, firma=ortak)
  Ortaklara Odenen/Murat eski borc -> kasa(GIDER, firma=ortak)
  Mal Alimi      -> hareketler(ALIS)
  banka sheet    -> ayni kurallar, banka_hesap_id=BANKA
  Odeme Plani    -> gelir_gider vadeli (Borc=GIDER, Alacak=GELIR)
  ortak maas     -> ortak cari odemesi (kasa GIDER firma=ortak)
"""
import os
import re
import sys
import uuid
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import psycopg2
from openpyxl import load_workbook
from db import DB_CONFIG

PATH = os.environ.get('XLSX_PATH', r"C:\Users\ykahm\OneDrive\Masaüstü\Kasa defteri.xlsx")
SCHEMA = "t_6"
COMMIT = '--commit' in sys.argv
NOW = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

URUN_SATIS = ('MOTO-PARCA', 'MOTOSİKLET PARÇASI')

# Rapor sayaclari
rapor = defaultdict(lambda: {'adet': 0, 'tutar': 0.0})
belirsiz = []


# ---------- yardimcilar ----------
def _norm(s):
    s = (s or '').lower()
    for a, b in [('ü','u'),('ö','o'),('ş','s'),('ı','i'),('ç','c'),('ğ','g'),('İ','i')]:
        s = s.replace(a, b)
    return s


def ortak_kod(aciklama):
    a = _norm(aciklama)
    if 'huseyin' in a:
        return 'ORT-HUSEYIN', 'Hüseyin Eyitutun'
    if 'ozkan' in a:
        return 'ORT-OZKAN', 'Özkan Bulut'
    if 'murat' in a:
        return 'ORT-MURAT', 'Murat Baykal'
    return None, None


KK_KARTLAR = {
    'ziraat': 'Ziraat Kredi Kartı',
    'akbank': 'Akbank Kredi Kartı',
    'qnb': 'QNB Kredi Kartı',
    'kuveyt': 'Kuveyttürk Kredi Kartı',
    'bonus': 'Bonuscard',
    'ykb': 'Yapı Kredi Kartı',
    'yapi kredi': 'Yapı Kredi Kartı',
}


def kk_kart_ad(aciklama):
    a = _norm(aciklama)
    for key, ad in KK_KARTLAR.items():
        if key in a:
            return ad
    return 'Diğer Kredi Kartı'


# Bilinen isci isimleri -> kanonik ad (tek personel karti)
ISCI_MAP = [
    ('habil', 'Habil Başer'),
    ('aslan', 'Aslan Göçer'),
    ('emre', 'Emre (Çalamacı)'),
    ('hamza', 'Hamza'),
    ('suleyman', 'Süleyman'),
    ('durmus', 'Durmuş'),
]


def isci_ad(aciklama):
    a = _norm(aciklama)
    for key, ad in ISCI_MAP:
        if key in a:
            return ad
    return 'İŞÇİLER (GENEL)'


def parse_tr_tutar(text):
    """TR formatli tutarlari aciklamadan cek: '4.611 tl', '1.313.451,25'."""
    if not text:
        return None
    cands = re.findall(r'\d{1,3}(?:\.\d{3})+(?:,\d+)?|\d+(?:,\d+)?', text)
    vals = []
    for s in cands:
        v = s.replace('.', '').replace(',', '.')
        try:
            f = float(v)
            if f >= 1:
                vals.append(f)
        except ValueError:
            pass
    return max(vals) if vals else None


def num(v):
    """Hucre degerini float'a cevir (bos -> 0)."""
    if v is None or str(v).strip() == '':
        return 0.0
    try:
        return float(str(v).replace(',', '.'))
    except ValueError:
        return 0.0


def tarih_str(v):
    if v is None:
        return ''
    s = str(v)
    return s.split()[0] if ' ' in s else s


def add(kategori, tutar):
    rapor[kategori]['adet'] += 1
    rapor[kategori]['tutar'] += float(tutar or 0)


# ---------- ana ----------
def main():
    mode = "COMMIT (GERCEK YAZMA)" if COMMIT else "DRY-RUN (sadece rapor)"
    print(f"{'='*70}\nMIGRATION t_6 — MOD: {mode}\n{'='*70}\n")

    wb = load_workbook(PATH, data_only=True)
    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(f"SET search_path TO {SCHEMA}, public")

    # Master kayit id cache
    banka_ids = {}      # ad -> id
    ortak_var = set()   # kod
    personel_ids = {}   # ad -> id

    def get_or_create_banka(ad, tip='BANKA'):
        if ad in banka_ids:
            return banka_ids[ad]
        cur.execute("SELECT id FROM banka_hesaplari WHERE ad=%s", (ad,))
        r = cur.fetchone()
        if r:
            banka_ids[ad] = r[0]
            return r[0]
        if COMMIT:
            cur.execute(
                "INSERT INTO banka_hesaplari (ad, tip, acilis_bakiye, aktif, created_at) "
                "VALUES (%s,%s,0,1,%s) RETURNING id", (ad, tip, NOW))
            bid = cur.fetchone()[0]
        else:
            bid = -len(banka_ids) - 1  # dry-run placeholder
        banka_ids[ad] = bid
        return bid

    def ensure_ortak(kod, ad):
        if kod in ortak_var:
            return
        cur.execute("SELECT kod FROM firmalar WHERE kod=%s", (kod,))
        if cur.fetchone():
            ortak_var.add(kod)
            return
        if COMMIT:
            cur.execute(
                "INSERT INTO firmalar (kod, ad, is_alani, aktif) VALUES (%s,%s,'ORTAK',1)", (kod, ad))
        ortak_var.add(kod)

    def get_or_create_personel(ad):
        if ad in personel_ids:
            return personel_ids[ad]
        cur.execute("SELECT id FROM personel WHERE ad=%s", (ad,))
        r = cur.fetchone()
        if r:
            personel_ids[ad] = r[0]
            return r[0]
        if COMMIT:
            cur.execute(
                "INSERT INTO personel (ad, durum, created_at) VALUES (%s,'AKTIF',%s) RETURNING id", (ad, NOW))
            pid = cur.fetchone()[0]
        else:
            pid = -len(personel_ids) - 1
        personel_ids[ad] = pid
        return pid

    # ---- TEMIZLIK (sadece commit) ----
    if COMMIT:
        print("[temizlik] t_6 test verisi siliniyor...")
        for t in ('hareketler', 'kasa', 'gelir_gider', 'cekler', 'personel_hareket',
                  'personel_aylik', 'personel', 'cek_hareketleri', 'banka_hesaplari'):
            cur.execute(f"DELETE FROM {t}")
        cur.execute("DELETE FROM firmalar WHERE kod LIKE 'TED%' OR kod LIKE 'MUS%' OR kod LIKE 'ORT-%'")
        cur.execute("DELETE FROM urunler WHERE kod IN ('PE001','KRT01',%s)", (URUN_SATIS[0],))
        print("  test verisi temizlendi\n")

    # ---- URUN: MOTOSIKLET PARCASI ----
    if COMMIT:
        cur.execute(
            "INSERT INTO urunler (kod, ad, kategori, birim, aktif) VALUES (%s,%s,'MOTOSİKLET','KG',1) "
            "ON CONFLICT (kod) DO NOTHING", URUN_SATIS)

    def kasa_ins(tarih, firma_kod, firma_ad, tur, tutar, odeme, aciklama,
                 banka_hesap_id=None, transfer_id='', is_transfer=0):
        if COMMIT:
            cur.execute(
                "INSERT INTO kasa (tarih, firma_kod, firma_ad, tur, tutar, odeme_sekli, aciklama, "
                "banka_hesap_id, transfer_id, is_transfer, created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (tarih, firma_kod, firma_ad, tur, tutar, odeme, aciklama,
                 banka_hesap_id, transfer_id, is_transfer, NOW))

    def hareket_ins(tarih, firma_kod, firma_ad, tur, miktar, tutar, aciklama):
        if COMMIT:
            bf = (tutar / miktar) if miktar else 0
            cur.execute(
                "INSERT INTO hareketler (tarih, firma_kod, firma_ad, tur, urun_kod, urun_ad, "
                "miktar, birim_fiyat, toplam, kdvli_toplam, aciklama, created_at) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (tarih, firma_kod, firma_ad, tur, URUN_SATIS[0], URUN_SATIS[1],
                 miktar, bf, tutar, tutar, aciklama, NOW))

    def gg_ins(tarih, tur, kategori, aciklama, tutar, odeme, firma_kod='', firma_ad='',
               banka_hesap_id=None, odeme_durumu='ODENDI', vade=''):
        """gelir_gider + (ODENDI ise) bagli kasa kaydi. Returns gg_id (commit) / None."""
        if not COMMIT:
            return None
        cur.execute(
            "INSERT INTO gelir_gider (tarih, tur, kategori, aciklama, tutar, toplam, odeme_sekli, "
            "firma_kod, firma_ad, odeme_durumu, vade_tarih, created_at) "
            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (tarih, tur, kategori, aciklama, tutar, tutar, odeme, firma_kod, firma_ad,
             odeme_durumu, vade, NOW))
        gg_id = cur.fetchone()[0]
        if odeme_durumu == 'ODENDI':
            kasa_tur = 'GIDER' if tur == 'GIDER' else 'GELIR'
            cur.execute(
                "INSERT INTO kasa (tarih, firma_kod, firma_ad, tur, tutar, odeme_sekli, aciklama, "
                "gelir_gider_id, banka_hesap_id, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                (tarih, firma_kod, firma_ad, kasa_tur, tutar, odeme,
                 f"GG: {kategori}: {aciklama}"[:200], gg_id, banka_hesap_id, NOW))
        return gg_id

    # ============ KASA + BANKA SHEET ISLEME ============
    def islem_satir(r, kaynak_banka_id):
        """Tek bir kasa/banka satirini kategoriye gore isle.
        kaynak_banka_id: None (kasa sheet=nakit) veya BANKA hesap id (banka sheet)."""
        tarih = tarih_str(r[0])
        aciklama = str(r[1] or '').strip()
        giren = num(r[2])
        cikan = num(r[3])
        turu = str(r[5] or '').strip()
        tn = _norm(turu)  # normalize: buyuk/kucuk harf + TR karakter duyarsiz
        kg = num(r[7]) if len(r) > 7 else 0
        if not turu:
            return
        tutar = giren if giren else cikan

        # banka sheet 'Kasa' kategorisi = kasa sheet 'Banka' ile ayni transfer.
        # Cift saymamak icin banka sheet tarafini ATLA (kasa sheet'te islendi).
        if tn == 'kasa':
            return

        if tn == 'satis':
            # stoklu satis + para girisi (kaynak hesaba)
            if giren > 0:
                hareket_ins(tarih, '', '', 'SATIS', kg, giren, aciklama)
                kasa_ins(tarih, '', '', 'GELIR', giren,
                         'BANKA' if kaynak_banka_id else 'NAKIT', aciklama, kaynak_banka_id)
                add('Satış', giren)
        elif tn == 'masraf':
            if cikan > 0:
                gg_ins(tarih, 'GIDER', 'MASRAF', aciklama, cikan,
                       'BANKA' if kaynak_banka_id else 'NAKIT', banka_hesap_id=kaynak_banka_id)
                add('Masraf', cikan)
        elif tn == 'maas':
            # isci: personel + personel_hareket + gelir_gider(GIDER, ISCILIK)
            if cikan > 0:
                ad = isci_ad(aciklama)
                pid = get_or_create_personel(ad)
                gg_id = gg_ins(tarih, 'GIDER', 'İŞÇİLİK', f"{ad}: {aciklama}", cikan,
                               'BANKA' if kaynak_banka_id else 'NAKIT', banka_hesap_id=kaynak_banka_id)
                if COMMIT:
                    yil, ay = (tarih[:4], tarih[5:7]) if len(tarih) >= 7 else ('2026', '01')
                    cur.execute(
                        "INSERT INTO personel_hareket (personel_id, yil, ay, tur, tutar, tarih, aciklama, gelir_gider_id, created_at) "
                        "VALUES (%s,%s,%s,'ODEME',%s,%s,%s,%s,%s)",
                        (pid, int(yil), int(ay), cikan, tarih, aciklama, gg_id, NOW))
                add('Maaş (işçi/personel)', cikan)
        elif tn == 'banka':
            # transfer nakit<->BANKA
            bid = get_or_create_banka('BANKA')
            tid = uuid.uuid4().hex
            if cikan > 0:  # bankaya yatan: nakit(-) -> banka(+)
                kasa_ins(tarih, '', '', 'GIDER', cikan, 'NAKIT', aciklama, None, tid, 1)
                kasa_ins(tarih, '', '', 'GELIR', cikan, 'BANKA', aciklama, bid, tid, 1)
                add('Banka transfer (nakit->banka)', cikan)
            elif giren > 0:  # bankadan cekilen: banka(-) -> nakit(+)
                kasa_ins(tarih, '', '', 'GIDER', giren, 'BANKA', aciklama, bid, tid, 1)
                kasa_ins(tarih, '', '', 'GELIR', giren, 'NAKIT', aciklama, None, tid, 1)
                add('Banka transfer (banka->nakit)', giren)
        elif tn == 'kk harcamasi':
            kart = kk_kart_ad(aciklama)
            kkid = get_or_create_banka(kart, tip='KREDI_KARTI')
            t = cikan if cikan else (giren if giren else parse_tr_tutar(aciklama))
            if t:
                # KK harcamasi = kredi kartindan gider
                gg_ins(tarih, 'GIDER', 'KK HARCAMASI', aciklama, t, 'BANKA', banka_hesap_id=kkid)
                add(f'KK: {kart}', t)
            else:
                belirsiz.append(f"KK tutar bulunamadi: {tarih} | {aciklama}")
        elif tn == 'ortaktan alinan':
            kod, ad = ortak_kod(aciklama)
            if not kod:
                belirsiz.append(f"Ortak tespit edilemedi: {tarih} | {aciklama}")
                return
            ensure_ortak(kod, ad)
            kasa_ins(tarih, kod, ad, 'GELIR', giren or tutar,
                     'BANKA' if kaynak_banka_id else 'NAKIT', aciklama, kaynak_banka_id)
            add('Ortaktan Alınan', giren or tutar)
        elif tn in ('ortaklara odenen', 'murat eski borc'):
            kod, ad = ortak_kod(aciklama)
            if tn == 'murat eski borc':
                kod, ad = 'ORT-MURAT', 'Murat Baykal'
            if not kod:
                belirsiz.append(f"Ortak tespit edilemedi: {tarih} | {aciklama}")
                return
            ensure_ortak(kod, ad)
            kasa_ins(tarih, kod, ad, 'GIDER', cikan or tutar,
                     'BANKA' if kaynak_banka_id else 'NAKIT', aciklama, kaynak_banka_id)
            add('Ortaklara Ödenen' if tn == 'ortaklara odenen' else 'Murat eski borç', cikan or tutar)
        elif tn == 'mal alimi':
            # Mal alimi -> MOTOSIKLET PARCASI stok girisi (ALIS) + para cikisi
            if cikan > 0:
                hareket_ins(tarih, '', '', 'ALIS', kg, cikan, aciklama)
                kasa_ins(tarih, '', '', 'GIDER', cikan,
                         'BANKA' if kaynak_banka_id else 'NAKIT', aciklama, kaynak_banka_id)
                add('Mal Alımı', cikan)
        else:
            belirsiz.append(f"Bilinmeyen kategori '{turu}': {tarih} | {aciklama}")

    # kasa sheet (nakit)
    ws = wb['kasa']
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if r[5] and str(r[5]).strip():
            islem_satir(r, None)

    # banka sheet (banka_hesap_id = BANKA). Bu sheet kol: TARIH,ACIKLAMA,GIREN,CIKAN,KALAN,Turu,KILOGRAM
    bid_banka = get_or_create_banka('BANKA')
    ws2 = wb['banka']
    for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
        if r[5] and str(r[5]).strip():
            # banka sheet KILOGRAM kol=6, kasa'da kol=7 — uyumlu olsun diye pad
            rr = list(r) + [None]  # kg index 7 olsun (banka'da kg kol6 -> kaydir)
            rr2 = [r[0], r[1], r[2], r[3], r[4], r[5], None, (r[6] if len(r) > 6 else None)]
            islem_satir(rr2, bid_banka)

    # ============ ODEME PLANI ============
    wsp = wb['Ödeme Planı']
    for r in wsp.iter_rows(min_row=2, max_row=wsp.max_row, values_only=True):
        kisi = str(r[1] or '').strip()
        ba = str(r[2] or '').strip()   # Borç/Alacak
        tutar = num(r[3])
        if not kisi or tutar <= 0:
            continue
        tarih = tarih_str(r[0])
        odendi = str(r[5] or '').strip().upper() == 'ÖDENDİ' if len(r) > 5 else False
        durum = 'ODENDI' if odendi else 'ODENMEDI'
        if 'alacak' in _norm(ba):
            gg_ins(tarih, 'GELIR', 'ÖDEME PLANI', kisi, tutar, 'NAKIT' if odendi else '',
                   odeme_durumu=durum, vade=tarih if not odendi else '')
            add('Ödeme Planı (Alacak)', tutar)
        else:
            gg_ins(tarih, 'GIDER', 'ÖDEME PLANI', kisi, tutar, 'NAKIT' if odendi else '',
                   odeme_durumu=durum, vade=tarih if not odendi else '')
            add('Ödeme Planı (Borç)', tutar)

    # ============ ORTAK MAAS TAKIBI ============
    # 3 blok: Murat(0-2), Ozkan(4-6), Huseyin(8-10) -> tarih,tutar,aciklama
    wsm = wb['ortak maaş takibi']
    bloklar = [(0, 'ORT-MURAT', 'Murat Baykal'), (4, 'ORT-OZKAN', 'Özkan Bulut'),
               (8, 'ORT-HUSEYIN', 'Hüseyin Eyitutun')]
    for r in wsm.iter_rows(min_row=2, max_row=wsm.max_row, values_only=True):
        for col, kod, ad in bloklar:
            if len(r) > col + 1 and r[col] and num(r[col + 1]) > 0:
                tarih = tarih_str(r[col])
                tutar = num(r[col + 1])
                acik = str(r[col + 2] or 'Ortak ücret').strip() if len(r) > col + 2 else 'Ortak ücret'
                ensure_ortak(kod, ad)
                kasa_ins(tarih, kod, ad, 'GIDER', tutar, 'NAKIT', f"Ortak ücret: {acik}")
                add('Ortak ücret', tutar)

    # ---- RAPOR ----
    print("KATEGORI ESLESMELERI:")
    toplam = 0.0
    for k in sorted(rapor.keys()):
        d = rapor[k]
        print(f"  {d['adet']:>4} kayıt | {d['tutar']:>14,.2f} TL | {k}")
        toplam += d['tutar']
    print(f"  {'-'*50}")
    print(f"  TOPLAM islenen tutar: {toplam:,.2f} TL")

    print(f"\nMASTER KAYITLAR:")
    print(f"  Banka/Kart hesaplari: {sorted(banka_ids.keys())}")
    print(f"  Ortak cariler       : {sorted(ortak_var)}")
    print(f"  Personel            : {sorted(personel_ids.keys())}")

    if belirsiz:
        print(f"\n⚠️ BELIRSIZ / ELLE BAKILACAK ({len(belirsiz)}):")
        for b in belirsiz[:40]:
            print(f"  - {b}")

    if COMMIT:
        conn.commit()
        print("\n✅ COMMIT edildi — veriler t_6'ya yazildi.")
    else:
        conn.rollback()
        print("\n(DRY-RUN — hicbir sey yazilmadi. Gercek icin: --commit)")
    conn.close()


if __name__ == "__main__":
    main()
