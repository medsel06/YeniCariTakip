# -*- coding: utf-8 -*-
"""t_6 MIGRATION v2 — muhasebe denetimi sonrasi DOGRU aktarim.

KULLANIM:
    python tools/migrate_excel_t6_v2.py            # DRY-RUN
    python tools/migrate_excel_t6_v2.py --commit   # GERCEK (once t_6 islem verisini temizler)

v1'e gore DUZELTMELER (Codex + muhasebe denetimi):
  1. Veresiye/musteri satislari -> musteri cari (Ismail/Emre/Ozcan), satis+tahsilat cariye
  2. Karttan NAKIT cekme -> nakit/banka GIRIS + kart borcu (v1 sadece gider yaziyordu)
  3. Odeme Plani -> odeme_takibi tablosu (GIDER DEGIL);
     kart kalemleri KART kaynagi ve ilgili kart hesabi ile takip edilir
  4. ortak maas takibi sheet -> ATLANIR (kasa/banka 'Ortaklara Odenen' ile cift)
  5. Acilis stok fisi -> stok eksi kapatilir (Excel oncesi donem)
  6. Kullanicinin F001-F009 cari kartlari KORUNUR (temizlikte silinmez)
  7. KDV brut (on muhasebe)
"""
import os, re, sys, uuid, unicodedata
from datetime import datetime
from collections import defaultdict
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import psycopg2
from openpyxl import load_workbook
from db import DB_CONFIG

PATH = os.environ.get('XLSX_PATH', r"C:\Users\ykahm\Downloads\Kasa defteri.xlsx")
SCHEMA = "t_6"
COMMIT = '--commit' in sys.argv
NOW = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
URUN = ('MOTO-PARCA', 'MOTOSİKLET PARÇASI')

rapor = defaultdict(lambda: {'adet': 0, 'tutar': 0.0})
notlar = []


def _norm(s):
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(ch for ch in s if not unicodedata.combining(ch)).lower()
    for a, b in [('ı','i'), ('İ','i')]:
        s = s.replace(a, b)
    return s


def num(v):
    if v is None or str(v).strip() == '':
        return 0.0
    try:
        return float(str(v).replace(',', '.'))
    except ValueError:
        return 0.0


def tarih_str(v):
    if v is None:
        return ''
    s = str(v)
    return s.split()[0] if ' ' in s else s


def add(k, t):
    rapor[k]['adet'] += 1
    rapor[k]['tutar'] += float(t or 0)


# --- musteri cari tespiti (satis) ---
MUSTERI = [('ismail altiparmak', 'CARI-ISMAIL', 'İsmail Altıparmak'),
           ('ismail', 'CARI-ISMAIL', 'İsmail Altıparmak'),
           ('emre', 'CARI-EMRE', 'Emre'),
           ('ozcan', 'CARI-OZCAN', 'Özcan'),
           ('ozcan a', 'CARI-OZCAN', 'Özcan')]


def musteri_bul(aciklama):
    a = _norm(aciklama)
    for key, kod, ad in MUSTERI:
        if key in a:
            return kod, ad
    return None, None


# --- urun tespiti (Huseyin'in URN-* kartlarina esle) ---
# Sira onemli: spesifik once. Eslesmezse MOTO-PARCA (genel).
URUN_ESLEME = [
    ('aku', ('URN-006', 'Akü')),
    ('kelle', ('URN-001', 'Kelle Motor')),
    ('hurda demir', ('URN-019', 'Hurda Demir')),
    ('koruma demir', ('URN-019', 'Hurda Demir')),
    ('kablo', ('URN-014', 'Kablo')),
    ('eksoz', ('URN-009', 'Eksoz')),
    ('ekzoz', ('URN-009', 'Eksoz')),
    ('jant', ('URN-021', 'Lastik-Jant')),
    ('lastik', ('URN-021', 'Lastik-Jant')),
    ('bakir', ('URN-002', 'Bakır')),
    ('aluminyum', ('URN-003', 'Alüminyum')),
    ('pirinc', ('URN-007', 'Pirinç')),
    ('dinamo', ('URN-008', 'Dinamo')),
    ('salter', ('URN-013', 'Şalter')),
    ('balans', ('URN-016', 'Balans')),
    ('buzdolabi', ('URN-017', 'Buzdolabı Motoru')),
    ('amortisor', ('URN-022', 'Yedek Parça')),
    ('on takim', ('URN-022', 'Yedek Parça')),
    ('motor', ('URN-004', 'Tüm Motosiklet')),
    ('cross', ('URN-024', 'Motosiklet Parçaları')),
    ('honda', ('URN-024', 'Motosiklet Parçaları')),
    ('parca', ('URN-024', 'Motosiklet Parçaları')),
]


def urun_bul(aciklama):
    a = _norm(aciklama)
    for key, (kod, ad) in URUN_ESLEME:
        if key in a:
            return kod, ad
    return URUN[0], URUN[1]


def ortak_kod(aciklama):
    a = _norm(aciklama)
    if 'huseyin' in a: return 'F009', 'Hüseyin Eyitutun'
    if 'ozkan' in a:   return 'F007', 'Özkan Bulut'
    if 'murat' in a:   return 'F008', 'Murat Baykal'
    return None, None


KK_KARTLAR = {'ziraat': 'Ziraat Kredi Kartı', 'akbank': 'Akbank Kredi Kartı',
              'qnb': 'QNB Kredi Kartı', 'kuveyt': 'Kuveyttürk Kredi Kartı',
              'bonus': 'Bonuscard', 'ykb': 'Yapı Kredi Kartı', 'yapi kredi': 'Yapı Kredi Kartı'}


def kk_kart_ad(aciklama):
    a = _norm(aciklama)
    for key, ad in KK_KARTLAR.items():
        if key in a:
            return ad
    return 'Diğer Kredi Kartı'


ISCI_MAP = [('habil', 'Habil Başer'), ('aslan', 'Aslan Göçer'), ('emre', 'Emre (Çalamacı)'),
            ('hamza', 'Hamza'), ('suleyman', 'Süleyman'), ('durmus', 'Durmuş')]


def isci_ad(aciklama):
    a = _norm(aciklama)
    for key, ad in ISCI_MAP:
        if key in a:
            return ad
    return 'İŞÇİLER (GENEL)'


def main():
    print(f"{'='*70}\nMIGRATION v2 t_6 — {'COMMIT' if COMMIT else 'DRY-RUN'}\n{'='*70}\n")
    print(f"Excel kaynak: {PATH}\n")
    wb = load_workbook(PATH, data_only=True)
    conn = psycopg2.connect(**DB_CONFIG); conn.autocommit = False
    cur = conn.cursor()
    cur.execute(f"SET search_path TO {SCHEMA}, public")

    banka_ids, banka_ad_by_id, cari_var, personel_ids = {}, {}, set(), {}
    tahmini_bakiye = defaultdict(float)
    toplam_satis_kg = [0.0]
    acilis_maliyet = [0.0]
    urun_satis = defaultdict(lambda: {'kg': 0.0, 'tutar': 0.0, 'ad': ''})  # urun bazli acilis icin

    def banka_id(ad, tip='BANKA'):
        if ad in banka_ids: return banka_ids[ad]
        cur.execute("SELECT id FROM banka_hesaplari WHERE ad=%s", (ad,))
        r = cur.fetchone()
        if r:
            banka_ids[ad] = r[0]
            banka_ad_by_id[r[0]] = ad
            return r[0]
        if COMMIT:
            cur.execute("INSERT INTO banka_hesaplari (ad,tip,acilis_bakiye,aktif,created_at) VALUES (%s,%s,0,1,%s) RETURNING id", (ad, tip, NOW))
            banka_ids[ad] = cur.fetchone()[0]
        else:
            banka_ids[ad] = -len(banka_ids) - 1
        banka_ad_by_id[banka_ids[ad]] = ad
        return banka_ids[ad]

    def ensure_cari(kod, ad, alan=''):
        if kod in cari_var: return
        cur.execute("SELECT kod FROM firmalar WHERE kod=%s", (kod,))
        if cur.fetchone(): cari_var.add(kod); return
        if COMMIT:
            cur.execute("INSERT INTO firmalar (kod,ad,is_alani,aktif) VALUES (%s,%s,%s,1)", (kod, ad, alan))
        cari_var.add(kod)

    def personel_id(ad):
        if ad in personel_ids: return personel_ids[ad]
        cur.execute("SELECT id FROM personel WHERE ad=%s", (ad,))
        r = cur.fetchone()
        if r: personel_ids[ad] = r[0]; return r[0]
        if COMMIT:
            cur.execute("INSERT INTO personel (ad,durum,created_at) VALUES (%s,'AKTIF',%s) RETURNING id", (ad, NOW))
            personel_ids[ad] = cur.fetchone()[0]
        else:
            personel_ids[ad] = -len(personel_ids) - 1
        return personel_ids[ad]

    # ---- TEMIZLIK (kullanicinin F00X carileri KORUNUR) ----
    if COMMIT:
        for t in ('hareketler','kasa','gelir_gider','cek_hareketleri','cekler',
                  'personel_hareket','personel_aylik','personel','odeme_takibi','banka_hesaplari'):
            cur.execute(f"DELETE FROM {t}")
        # Sadece migration uretimi carileri sil (F00X kullanici kartlari kalir)
        cur.execute("DELETE FROM firmalar WHERE kod LIKE 'TED%' OR kod LIKE 'MUS%' OR kod LIKE 'ORT-%' OR kod LIKE 'CARI-%'")
        cur.execute("INSERT INTO urunler (kod,ad,kategori,birim,aktif) VALUES (%s,%s,'MOTOSİKLET','KG',1) ON CONFLICT (kod) DO NOTHING", URUN)
        print("[temizlik] islem verisi silindi, F00X cariler korundu\n")

    def kasa_takip(tur, tutar, bhid=None):
        hesap = 'NAKIT' if bhid is None else banka_ad_by_id.get(bhid, f'HESAP:{bhid}')
        tahmini_bakiye[hesap] += float(tutar or 0) if tur == 'GELIR' else -float(tutar or 0)

    def kasa_ins(tarih, fk, fa, tur, tutar, odeme, acik, bhid=None, tid='', istr=0):
        kasa_takip(tur, tutar, bhid)
        if COMMIT:
            cur.execute("INSERT INTO kasa (tarih,firma_kod,firma_ad,tur,tutar,odeme_sekli,aciklama,banka_hesap_id,transfer_id,is_transfer,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (tarih, fk, fa, tur, tutar, odeme, acik[:200], bhid, tid, istr, NOW))

    def hareket_ins(tarih, fk, fa, tur, miktar, tutar, acik, urun_kod=None, urun_ad=None):
        uk = urun_kod or URUN[0]; ua = urun_ad or URUN[1]
        if COMMIT:
            bf = (tutar / miktar) if miktar else 0
            cur.execute("INSERT INTO hareketler (tarih,firma_kod,firma_ad,tur,urun_kod,urun_ad,miktar,birim_fiyat,toplam,kdvli_toplam,aciklama,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (tarih, fk, fa, tur, uk, ua, miktar, bf, tutar, tutar, acik[:200], NOW))

    def gg_ins(tarih, tur, kat, acik, tutar, odeme, fk='', fa='', bhid=None, durum='ODENDI', kasa_yarat=True):
        if durum == 'ODENDI' and kasa_yarat:
            kasa_takip('GIDER' if tur == 'GIDER' else 'GELIR', tutar, bhid)
        if not COMMIT: return None
        cur.execute("INSERT INTO gelir_gider (tarih,tur,kategori,aciklama,tutar,toplam,odeme_sekli,firma_kod,firma_ad,odeme_durumu,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                    (tarih, tur, kat, acik[:200], tutar, tutar, odeme, fk, fa, durum, NOW))
        gid = cur.fetchone()[0]
        if durum == 'ODENDI' and kasa_yarat:
            kt = 'GIDER' if tur == 'GIDER' else 'GELIR'
            cur.execute("INSERT INTO kasa (tarih,firma_kod,firma_ad,tur,tutar,odeme_sekli,aciklama,gelir_gider_id,banka_hesap_id,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (tarih, fk, fa, kt, tutar, odeme, f"GG: {kat}: {acik}"[:200], gid, bhid, NOW))
        return gid

    def odeme_takibi_ins(tip, kaynak, fk, fa, acik, tutar, vade, durum='ACIK', bhid=None):
        if COMMIT:
            cur.execute("INSERT INTO odeme_takibi (tip,kaynak,firma_kod,firma_ad,banka_hesap_id,aciklama,tutar,odenen,vade_tarih,durum,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,0,%s,%s,%s)",
                        (tip, kaynak, fk, fa, bhid, acik[:200], tutar, vade, durum, NOW))

    # ========== KASA + BANKA SHEET ==========
    def islem(r, kaynak_bhid):
        tarih = tarih_str(r[0]); acik = str(r[1] or '').strip()
        giren = num(r[2]); cikan = num(r[3]); turu = str(r[5] or '').strip()
        tn = _norm(turu); kg = num(r[7]) if len(r) > 7 else 0
        odeme_tarih = tarih_str(r[6]) if len(r) > 6 and r[6] else ''
        if not turu: return
        odeme_sekli = 'BANKA' if kaynak_bhid else 'NAKIT'

        if tn == 'kasa':
            return  # banka sheet transfer -> kasa sheet 'Banka'da islendi
        if tn == 'satis':
            if giren <= 0: return
            mk, ma = musteri_bul(acik)
            uk, ua = urun_bul(acik)
            hareket_ins(tarih, mk or '', ma or '', 'SATIS', kg, giren, acik, uk, ua)
            toplam_satis_kg[0] += kg
            urun_satis[uk]['kg'] += kg
            urun_satis[uk]['tutar'] += giren
            urun_satis[uk]['ad'] = ua
            if mk:
                ensure_cari(mk, ma, 'MUSTERI')
                # satis cariye alacak dogurur; tahsil edilen (giren) ayni cariye tahsilat
                kasa_ins(tarih, mk, ma, 'GELIR', giren, odeme_sekli, f"Tahsilat: {acik}", kaynak_bhid)
                add('Satış (müşteri cari)', giren)
            else:
                kasa_ins(tarih, '', '', 'GELIR', giren, odeme_sekli, acik, kaynak_bhid)
                add('Satış (peşin/anonim)', giren)
        elif tn == 'masraf':
            if cikan > 0:
                gg_ins(tarih, 'GIDER', 'MASRAF', acik, cikan, odeme_sekli, bhid=kaynak_bhid)
                add('Masraf', cikan)
        elif tn == 'maas':
            if cikan > 0:
                ad = isci_ad(acik); pid = personel_id(ad)
                gid = gg_ins(tarih, 'GIDER', 'İŞÇİLİK', f"{ad}: {acik}", cikan, odeme_sekli, bhid=kaynak_bhid)
                if COMMIT:
                    yil, ay = (tarih[:4], tarih[5:7]) if len(tarih) >= 7 else ('2026', '01')
                    cur.execute("INSERT INTO personel_hareket (personel_id,yil,ay,tur,tutar,tarih,aciklama,gelir_gider_id,created_at) VALUES (%s,%s,%s,'ODEME',%s,%s,%s,%s,%s)",
                                (pid, int(yil), int(ay), cikan, tarih, acik, gid, NOW))
                add('Maaş (işçi/personel)', cikan)
        elif tn == 'banka':
            bid = banka_id('BANKA'); tid = uuid.uuid4().hex
            if cikan > 0:  # bankaya yatan
                kasa_ins(tarih, '', '', 'GIDER', cikan, 'NAKIT', acik, None, tid, 1)
                kasa_ins(tarih, '', '', 'GELIR', cikan, 'BANKA', acik, bid, tid, 1)
                add('Transfer nakit->banka', cikan)
            elif giren > 0:  # bankadan cekilen
                kasa_ins(tarih, '', '', 'GIDER', giren, 'BANKA', acik, bid, tid, 1)
                kasa_ins(tarih, '', '', 'GELIR', giren, 'NAKIT', acik, None, tid, 1)
                add('Transfer banka->nakit', giren)
        elif tn == 'kk harcamasi':
            kart = kk_kart_ad(acik)
            a = _norm(acik)
            if 'cekilen' in a:  # KARTTAN NAKIT CEKME -> nakit giris + kart borcu
                brut = _parse_tutar(acik) or cikan or giren
                net = giren or cikan or brut
                # banka sheet'te kart adi yazmiyor ("Kredi kartindan cekilen") -> tutar
                # bazli kart tespiti (Codex): 500.000 Bonuscard, 375.000 Yapi Kredi
                if kart == 'Diğer Kredi Kartı' and brut:
                    if abs(float(brut) - 500000) < 1:
                        kart = 'Bonuscard'
                    elif abs(float(brut) - 375000) < 1:
                        kart = 'Yapı Kredi Kartı'
                kkid = banka_id(kart, 'KREDI_KARTI')
                if brut:
                    kasa_ins(tarih, '', '', 'GIDER', brut, 'BANKA', f"Karttan çekilen brüt borç: {acik}", kkid)
                    kasa_ins(tarih, '', '', 'GELIR', net, odeme_sekli, f"Karttan çekilen net giriş: {acik}", kaynak_bhid)
                    komisyon = max(float(brut or 0) - float(net or 0), 0)
                    if komisyon:
                        gg_ins(tarih, 'GIDER', 'KART NAKİT AVANS KOMİSYON', acik, komisyon, 'BANKA',
                               bhid=kkid, kasa_yarat=False)
                        add('Kart nakit çekme komisyonu', komisyon)
                    add('Karttan nakit çekme (brüt kart borcu)', brut)
                    add('Karttan nakit çekme (net giriş)', net)
            else:  # kart harcamasi = gider + kart borcu
                kkid = banka_id(kart, 'KREDI_KARTI')
                t = cikan or giren or _parse_tutar(acik)
                if t:
                    gg_ins(tarih, 'GIDER', 'KK HARCAMASI', acik, t, 'BANKA', bhid=kkid)
                    add(f'KK: {kart}', t)
                else:
                    notlar.append(f"KK tutar yok: {tarih} | {acik}")
        elif tn == 'ortaktan alinan':
            kod, ad = ortak_kod(acik)
            if not kod: notlar.append(f"Ortak yok: {acik}"); return
            ensure_cari(kod, ad, 'ORTAK')
            kasa_ins(tarih, kod, ad, 'GELIR', giren or (giren or cikan), odeme_sekli, acik, kaynak_bhid)
            add('Ortaktan Alınan', giren)
        elif tn in ('ortaklara odenen', 'murat eski borc'):
            kod, ad = ortak_kod(acik)
            if tn == 'murat eski borc': kod, ad = 'F008', 'Murat Baykal'
            if not kod: notlar.append(f"Ortak yok: {acik}"); return
            ensure_cari(kod, ad, 'ORTAK')
            kasa_ins(tarih, kod, ad, 'GIDER', cikan, odeme_sekli, acik, kaynak_bhid)
            add('Ortaklara Ödenen' if tn == 'ortaklara odenen' else 'Murat eski borç', cikan)
        elif tn == 'mal alimi':
            if cikan > 0:
                hareket_ins(tarih, '', '', 'ALIS', kg, cikan, acik)
                acilis_maliyet[0] += cikan
                kasa_ins(tarih, '', '', 'GIDER', cikan, odeme_sekli, acik, kaynak_bhid)
                add('Mal Alımı', cikan)

    def _parse_tutar(text):
        cands = re.findall(r'\d{1,3}(?:\.\d{3})+(?:,\d+)?|\d+(?:,\d+)?', text or '')
        vals = []
        for s in cands:
            try:
                f = float(s.replace('.', '').replace(',', '.'))
                if f >= 1: vals.append(f)
            except ValueError: pass
        return max(vals) if vals else None

    ws = wb['kasa']
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if r[5] and str(r[5]).strip(): islem(r, None)
    bid_banka = banka_id('BANKA')
    ws2 = wb['banka']
    for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row, values_only=True):
        if r[5] and str(r[5]).strip():
            rr = [r[0], r[1], r[2], r[3], r[4], r[5], None, (r[6] if len(r) > 6 else None)]
            islem(rr, bid_banka)

    # ========== ODEME PLANI -> odeme_takibi ==========
    KART_ISARET = ['ykb','bonus','kuveyt','qnb','akbank','ziraat kart','market harcama','cayci',
                   'somun','ayyildiz','sanayi tupu','yemek ucret','spiral','cekilen','miles']
    CARI_ESLEME = [('aybimas',('CARI-AYBIMAS','Aybimaş AŞ')), ('mke',('F006','MKE AŞ')),
                   ('dilek',('CARI-DILEK','Dilek Geri Dönüşüm')), ('oknal',('CARI-OKNAL','Oknal Gaz')),
                   ('ismail',('CARI-ISMAIL','İsmail Altıparmak'))]
    wsp = wb['Ödeme Planı']
    for r in wsp.iter_rows(min_row=2, max_row=wsp.max_row, values_only=True):
        kisi = str(r[1] or '').strip(); tutar = num(r[3])
        if not kisi or tutar <= 0: continue
        tarih = tarih_str(r[0]); ba = _norm(str(r[2] or ''))
        a = _norm(kisi)
        kod, ad = None, None
        kart = kk_kart_ad(kisi)
        kart_mi = kart != 'Diğer Kredi Kartı' or any(x in a for x in KART_ISARET)
        if kart_mi:
            kkid = banka_id(kart, 'KREDI_KARTI')
            odeme_takibi_ins('BORC', 'KART', '', kart, kisi, tutar, tarih, bhid=kkid)
            add('Ödeme Takibi (KART)', tutar)
            continue
        for key, (k, ad2) in CARI_ESLEME:
            if key in a: kod, ad = k, ad2; break
        if not kod:
            kod, ad = 'CARI-' + re.sub(r'[^A-Z0-9]', '', _norm(kisi).upper())[:14], kisi
        ensure_cari(kod, ad)
        tip = 'ALACAK' if 'alacak' in ba else 'BORC'
        odeme_takibi_ins(tip, 'CARI', kod, ad, kisi, tutar, tarih)
        add(f'Ödeme Takibi ({tip})', tutar)

    # ========== ACILIS STOK FISI (URUN BAZLI — her urunde stok dengelenir) ==========
    # Her urun icin satilan kadar acilis girisi; maliyet = o urunun satis tutari
    # (gercek maliyet bilinmedigi icin maliyet~satis, acilis kari 0). Stok her urunde ~0.
    if urun_satis:
        for uk, d in sorted(urun_satis.items()):
            if d['kg'] > 0:
                hareket_ins('2026-04-15', '', '', 'ALIS', d['kg'], d['tutar'],
                            f"AÇILIŞ STOĞU - {d['ad']} (Excel öncesi)", uk, d['ad'])
                add('Açılış stok fişi (ürün bazlı)', d['tutar'])
        notlar.append(f"Açılış stok fişi: {len(urun_satis)} ürün, toplam {toplam_satis_kg[0]:,.0f} kg "
                      f"(her ürün satılan kadar giriş, maliyet~satış)")

    notlar.append("ortak maaş takibi sheet ATLANDI (kasa/banka 'Ortaklara Ödenen' ile çift sayım)")

    # ========== KASA MUTABAKAT FARKI (Excel son kalan ile esitleme) ==========
    # Nakit kasa tahmini bakiyeyi Excel kasa defteri son KALAN'ina esitle.
    # P&L'ye girmez (kategori=MUTABAKAT, gelir_gider'e yazilmaz), cari/stok/kart etkilemez.
    EXCEL_NAKIT_KALAN = 159950.0
    fark = round(tahmini_bakiye.get('NAKIT', 0.0) - EXCEL_NAKIT_KALAN, 2)
    if abs(fark) > 0.01:
        tur_m = 'GIDER' if fark > 0 else 'GELIR'  # fazlaysa GIDER ile dus
        kasa_takip(tur_m, abs(fark), None)
        if COMMIT:
            cur.execute(
                "INSERT INTO kasa (tarih,firma_kod,firma_ad,tur,tutar,odeme_sekli,aciklama,kategori,created_at) "
                "VALUES (%s,'','',%s,%s,'NAKIT',%s,'MUTABAKAT',%s)",
                ('2026-04-15', tur_m, abs(fark),
                 'MİGRASYON KASA MUTABAKAT FARKI - Excel son kalan 159.950 TL ile eşitleme', NOW))
        add('MİGRASYON KASA MUTABAKAT FARKI', abs(fark))
        notlar.append(f"Kasa mutabakat farkı satırı: {tur_m} {abs(fark):,.2f} TL (nakit -> 159.950,00)")

    # ---- RAPOR ----
    print("KATEGORI:")
    for k in sorted(rapor.keys()):
        d = rapor[k]; print(f"  {d['adet']:>4} | {d['tutar']:>14,.2f} TL | {k}")
    print(f"\nBanka/Kart: {sorted(banka_ids.keys())}")
    print(f"Yeni cari : {sorted(cari_var)}")
    print(f"Personel  : {sorted(personel_ids.keys())}")
    print("\nTAHMINI HESAP BAKIYELERI (kasa tablosu etkisi):")
    for ad, bakiye in sorted(tahmini_bakiye.items()):
        print(f"  {ad:>24}: {bakiye:>14,.2f} TL")
    if notlar:
        print("\nNOTLAR:")
        for n in notlar: print(f"  - {n}")

    if COMMIT:
        conn.commit(); print("\n✅ COMMIT edildi.")
    else:
        conn.rollback(); print("\n(DRY-RUN — --commit ile uygula)")
    conn.close()


if __name__ == "__main__":
    main()
