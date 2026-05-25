# -*- coding: utf-8 -*-
"""Odeme Plani sheet'indeki benzersiz kisi/firma adlarini cikar."""
from openpyxl import load_workbook
import os
PATH = os.environ.get('XLSX_PATH', r"C:\Users\ykahm\OneDrive\Masaüstü\Kasa defteri.xlsx")
wb = load_workbook(PATH, data_only=True)
ws = wb['Ödeme Planı']
from collections import defaultdict
kisiler = defaultdict(lambda: {'adet': 0, 'borc': 0.0, 'alacak': 0.0})
for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    kisi = str(r[1] or '').strip()
    if not kisi:
        continue
    ba = str(r[2] or '').strip().lower()
    try:
        tutar = float(str(r[3]).replace(',', '.')) if r[3] else 0
    except Exception:
        tutar = 0
    k = kisiler[kisi]
    k['adet'] += 1
    if 'alacak' in ba:
        k['alacak'] += tutar
    else:
        k['borc'] += tutar
print(f"Benzersiz kisi/firma: {len(kisiler)}\n")
for kisi, d in sorted(kisiler.items(), key=lambda x: -(x[1]['borc'] + x[1]['alacak'])):
    print(f"  {d['adet']:>2}x | borc={d['borc']:>13,.2f} | alacak={d['alacak']:>11,.2f} | {kisi}")
