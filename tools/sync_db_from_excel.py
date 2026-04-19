import os
import re
import sys
import shutil
import sqlite3
import unicodedata
from datetime import datetime

from openpyxl import load_workbook


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "alse_muhasebe.db")

EXCLUDE_SHEETS = {
    "ÜRÜN LİSTESİ",
    "STOK TAKİP",
    "FİRMA LİSTESİ",
    "Hedef Sayfa",
    "ÇEK",
    "YENİ FİRMA",
    "YENİ FİRMA (2)",
    "YENİ FİRMA (3)",
}


def norm(text):
    text = str(text or "").strip().casefold()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    text = text.translate(str.maketrans({"ı": "i", "ş": "s", "ç": "c", "ö": "o", "ü": "u", "ğ": "g"}))
    text = "".join(ch for ch in text if ch.isalnum() or ch.isspace())
    return " ".join(text.split())


def find_sheet_by_norm(wb, wanted_norm):
    for s in wb.sheetnames:
        if norm(s) == wanted_norm:
            return s
    return None


def parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    if len(digits) < 7:
        return None
    if len(digits) > 8:
        digits = digits[-8:]
    if len(digits) == 8:
        d = int(digits[:2])
        m = int(digits[2:4])
        y = int(digits[4:8])
        try:
            return datetime(y, m, d).strftime("%Y-%m-%d")
        except ValueError:
            return None
    y = int(digits[-4:])
    left = digits[:-4]
    if len(left) == 3:
        cands = [(int(left[0]), int(left[1:])), (int(left[:2]), int(left[2:]))]
        for d, m in cands:
            try:
                return datetime(y, m, d).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def fnum(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    s = s.replace(" ", "")
    s = s.replace(".", "").replace(",", ".") if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", s) else s
    try:
        return float(s)
    except ValueError:
        return 0.0


def infer_payment_method(desc):
    t = norm(desc)
    if "cek" in t:
        return "CEK"
    if "havale" in t or "banka" in t or "eft" in t:
        return "HAVALE"
    return "NAKIT"


def resolve_excel_path(arg):
    if arg and os.path.exists(arg):
        return arg
    desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "Masaüstü")
    if os.path.isdir(desktop):
        for name in os.listdir(desktop):
            if name.upper().startswith("ALSE") and name.lower().endswith(".xlsm"):
                return os.path.join(desktop, name)
    raise FileNotFoundError("Excel dosyasi bulunamadi. Arguman olarak tam yolu verin.")


def backup_db():
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"alse_muhasebe_before_excel_sync_{ts}.db"
    backup_path = os.path.join(BASE_DIR, backup_name)
    shutil.copy2(DB_PATH, backup_path)
    return backup_path


def main():
    excel_path = resolve_excel_path(sys.argv[1] if len(sys.argv) > 1 else None)
    backup_path = backup_db()

    wb = load_workbook(excel_path, data_only=True, read_only=True)

    product_sheet = find_sheet_by_norm(wb, "urun listesi")
    cek_sheet = find_sheet_by_norm(wb, "cek")
    if not product_sheet:
        raise RuntimeError("Excel'de 'ÜRÜN LİSTESİ' sayfasi bulunamadi.")

    products = []
    product_map = {}
    auto_u_idx = 1

    ws_products = wb[product_sheet]
    for row in ws_products.iter_rows(min_row=4, max_col=3, values_only=True):
        ad = (row[0] or "").strip() if isinstance(row[0], str) else str(row[0] or "").strip()
        if not ad:
            continue
        kategori = (row[1] or "").strip() if isinstance(row[1], str) else str(row[1] or "").strip()
        birim = (row[2] or "KG").strip() if isinstance(row[2], str) else str(row[2] or "KG").strip()
        kod = f"U{len(products)+1:03d}"
        products.append({"kod": kod, "ad": ad, "kategori": kategori, "birim": birim or "KG"})
        product_map[norm(ad)] = kod

    firm_sheet_names = [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]
    firms = []
    firm_code_map = {}
    for idx, firm_name in enumerate(firm_sheet_names, start=1):
        kod = f"F{idx:03d}"
        firm_ad = firm_name.strip()
        firms.append({"kod": kod, "ad": firm_ad})
        firm_code_map[norm(firm_ad)] = kod

    def ensure_product(product_name):
        nonlocal auto_u_idx
        n = norm(product_name)
        if not n:
            return ""
        if n in product_map:
            return product_map[n]
        while True:
            kod = f"UX{auto_u_idx:03d}"
            auto_u_idx += 1
            if kod not in {p["kod"] for p in products}:
                break
        ad = str(product_name).strip()
        products.append({"kod": kod, "ad": ad, "kategori": "DİĞER", "birim": "KG"})
        product_map[n] = kod
        return kod

    def ensure_firm(firm_name):
        n = norm(firm_name)
        if n in firm_code_map:
            return firm_code_map[n]
        kod = f"F{len(firms)+1:03d}"
        ad = str(firm_name).strip()
        firms.append({"kod": kod, "ad": ad})
        firm_code_map[n] = kod
        return kod

    hareketler = []
    kasa_rows = []

    for sheet_name in firm_sheet_names:
        ws = wb[sheet_name]
        firma_ad = sheet_name.strip()
        firma_kod = ensure_firm(firma_ad)

        empty_run = 0
        for row in ws.iter_rows(min_row=5, max_col=23, values_only=True):
            has_left = (row[0] not in (None, "")) or (row[1] not in (None, "")) or (fnum(row[7]) != 0)
            has_right = (row[8] not in (None, "")) or (row[9] not in (None, "")) or (fnum(row[16]) != 0)
            has_pay = (row[19] not in (None, "")) or (row[20] not in (None, "")) or (fnum(row[21]) != 0)
            if not (has_left or has_right or has_pay):
                empty_run += 1
                if empty_run >= 120:
                    break
                continue
            empty_run = 0

            left_kdvli = fnum(row[7])
            if abs(left_kdvli) > 0:
                urun_ad = str(row[0] or "").strip()
                hareketler.append({
                    "tarih": parse_excel_date(row[1]),
                    "firma_kod": firma_kod,
                    "firma_ad": firma_ad,
                    "tur": "SATIS",
                    "urun_kod": ensure_product(urun_ad),
                    "urun_ad": urun_ad,
                    "miktar": fnum(row[2]),
                    "birim_fiyat": fnum(row[3]),
                    "toplam": fnum(row[4]) or left_kdvli,
                    "kdv_orani": fnum(row[5]),
                    "kdv_tutar": fnum(row[6]),
                    "kdvli_toplam": left_kdvli,
                })

            right_kdvli = fnum(row[16])
            if abs(right_kdvli) > 0:
                urun_ad = str(row[8] or "").strip()
                hareketler.append({
                    "tarih": parse_excel_date(row[9]),
                    "firma_kod": firma_kod,
                    "firma_ad": firma_ad,
                    "tur": "ALIS",
                    "urun_kod": ensure_product(urun_ad),
                    "urun_ad": urun_ad,
                    "miktar": fnum(row[10]),
                    "birim_fiyat": fnum(row[11]),
                    "toplam": fnum(row[12]) or right_kdvli,
                    "kdv_orani": fnum(row[13]),
                    "kdv_tutar": fnum(row[14]),
                    "kdvli_toplam": right_kdvli,
                })

            signed_payment = fnum(row[21])
            if abs(signed_payment) > 0:
                aciklama = str(row[20] or "").strip()
                kasa_rows.append({
                    "tarih": parse_excel_date(row[19]),
                    "firma_kod": firma_kod,
                    "firma_ad": firma_ad,
                    "tur": "GELIR" if signed_payment < 0 else "GIDER",
                    "tutar": abs(signed_payment),
                    "odeme_sekli": infer_payment_method(aciklama),
                    "aciklama": aciklama,
                })

    cek_rows = []
    if cek_sheet:
        ws_cek = wb[cek_sheet]
        cek_no = 1
        for row in ws_cek.iter_rows(min_row=6, max_col=6, values_only=True):
            firma_ad = str(row[3] or "").strip()
            tutar = fnum(row[5])
            if not firma_ad or abs(tutar) <= 0:
                continue
            firma_kod = ensure_firm(firma_ad)
            cek_turu = "ALINAN" if tutar > 0 else "VERILEN"
            cek_rows.append({
                "cek_no": f"CEK{cek_no:04d}",
                "firma_kod": firma_kod,
                "firma_ad": firma_ad,
                "kesim_tarih": None,
                "vade_tarih": parse_excel_date(row[4]),
                "tutar": abs(tutar),
                "tur": "ALACAK" if tutar > 0 else "BORÇ",
                "durum": "PORTFOYDE" if cek_turu == "ALINAN" else "KESILDI",
                "cek_turu": cek_turu,
            })
            cek_no += 1

    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("PRAGMA foreign_keys=OFF")
        conn.execute("BEGIN")

        conn.execute("DELETE FROM cek_hareketleri")
        conn.execute("DELETE FROM kasa")
        conn.execute("DELETE FROM hareketler")
        conn.execute("DELETE FROM cekler")
        conn.execute("DELETE FROM firmalar")
        conn.execute("DELETE FROM urunler")
        conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('hareketler','kasa','cekler','cek_hareketleri')")

        conn.executemany(
            """
            INSERT INTO urunler (kod, ad, kategori, birim) VALUES (?, ?, ?, ?)
            """,
            [(p["kod"], p["ad"], p["kategori"], p["birim"]) for p in products],
        )
        conn.executemany(
            """
            INSERT INTO firmalar (kod, ad, tel, adres, vkn_tckn, nace, is_alani, email, risk_limiti)
            VALUES (?, ?, '', '', '', '', '', '', 0)
            """,
            [(f["kod"], f["ad"]) for f in firms],
        )
        conn.executemany(
            """
            INSERT INTO hareketler (
                tarih, firma_kod, firma_ad, tur, urun_kod, urun_ad,
                miktar, birim_fiyat, toplam, kdv_orani, kdv_tutar, kdvli_toplam
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    h["tarih"], h["firma_kod"], h["firma_ad"], h["tur"], h["urun_kod"], h["urun_ad"],
                    h["miktar"], h["birim_fiyat"], h["toplam"], h["kdv_orani"], h["kdv_tutar"], h["kdvli_toplam"],
                )
                for h in hareketler
            ],
        )
        conn.executemany(
            """
            INSERT INTO kasa (tarih, firma_kod, firma_ad, tur, tutar, odeme_sekli, aciklama)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (k["tarih"], k["firma_kod"], k["firma_ad"], k["tur"], k["tutar"], k["odeme_sekli"], k["aciklama"])
                for k in kasa_rows
            ],
        )
        conn.executemany(
            """
            INSERT INTO cekler (
                cek_no, firma_kod, firma_ad, kesim_tarih, vade_tarih, tutar,
                tur, durum, cek_turu, kesideci, lehtar, ciro_firma_kod, ciro_firma_ad, tahsil_tarih, notlar
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, '', '', '', '', NULL, '')
            """,
            [
                (
                    c["cek_no"], c["firma_kod"], c["firma_ad"], c["kesim_tarih"], c["vade_tarih"],
                    c["tutar"], c["tur"], c["durum"], c["cek_turu"],
                )
                for c in cek_rows
            ],
        )

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    print(f"Excel: {excel_path}")
    print(f"Backup: {backup_path}")
    print(f"Urunler: {len(products)}")
    print(f"Firmalar: {len(firms)}")
    print(f"Hareketler: {len(hareketler)}")
    print(f"Kasa: {len(kasa_rows)}")
    print(f"Cekler: {len(cek_rows)}")


if __name__ == "__main__":
    main()
